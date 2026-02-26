import datetime
import os
import openpyxl
from typing import Dict, Any, List
from .base_evaluator import BaseEvaluator
from ..utils import load_dataset

def datetime_to_float(dt):
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_cell_value(v1, v2):
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        return False
    if v1 == v2:
        return True
    else:
        return False


def col_num2name(n):
    """ Convert a column number to an Excel column name """
    name = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    """ Convert an Excel column name to a column number """
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num


def parse_cell_range(range_str):
    """ Parse a range string like 'A1:AB12' """
    start_cell, end_cell = range_str.split(':')
    start_col, start_row = '', ''
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char
    
    end_col, end_row = '', ''
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return (col_name2num(start_col), int(start_row)), (col_name2num(end_col), int(end_row))


def generate_cell_names(range_str):
    """ Generate a list of all cell names in the specified range """
    if ':' not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [col_num2name(i) for i in range(start_col, end_col + 1)]
    cell_names = [f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)]
    return cell_names


def cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range):
    if sheet_name not in wb_proc:
        return False, f"worksheet '{sheet_name}' not found"
    
    # Check if sheet exists in GT (it should)
    if sheet_name not in wb_gt:
        # Fallback to first sheet if specified sheet not in GT? 
        # But logic says sheet_name comes from config, so it should exist in GT.
        # However, test logic had:
        # if '!' in sheet_cell_range: ... else: sheet_name = wb_gt.sheetnames[0]
        # We will assume sheet_name is correct for now.
        if len(wb_gt.sheetnames) > 0:
            ws_gt = wb_gt[wb_gt.sheetnames[0]]
        else:
            return False, "GT workbook has no sheets"
    else:
        ws_gt = wb_gt[sheet_name]
        
    ws_proc = wb_proc[sheet_name]

    cell_names = generate_cell_names(cell_range)

    for cell_name in cell_names:
        cell_gt = ws_gt[cell_name]
        cell_proc = ws_proc[cell_name]
        
        if not compare_cell_value(cell_gt.value, cell_proc.value):
            msg = f"Value difference at cell {cell_gt.coordinate}: ws_gt has {cell_gt.value}, ws_proc has {cell_proc.value}"
            return False, msg
        
    return True, ""


class FileGenerationEvaluator(BaseEvaluator):
    def evaluate_single(self, row: Dict[str, Any]) -> Dict[str, Any]:
        # This evaluator processes the whole dataset against one prediction file usually,
        # or expects 'pred_file' in row. 
        # But looking at auto_eval.py, evaluate_dataset is called.
        pass

    def evaluate_dataset(self, data_path: str, pred_path: str) -> List[Dict[str, Any]]:
        print(f"Loading file generation dataset from {data_path}...")
        data = load_dataset(data_path)
        
        print(f"Prediction file: {pred_path}")
        
        if not os.path.exists(pred_path):
            print(f"Error: Prediction file not found at {pred_path}")
            return []

        results = []
        total_score = 0
        
        # Open prediction workbook once if possible? 
        # But comparison might need fresh open or different options.
        # We will open it inside the loop or helper to be safe, but optimization would be opening once.
        # However, compare logic takes paths usually. 
        # Let's open workbooks here to avoid re-opening for every item if they share the file.
        # But items might refer to different GT files.
        # The pred_path is constant (one file).
        
        try:
            wb_proc = openpyxl.load_workbook(filename=pred_path, data_only=True)
        except Exception as e:
            print(f"Failed to open prediction file: {e}")
            return []

        for i, item in enumerate(data):
            gt_file_rel = item.get('gt_file')
            # Resolve gt_file path relative to data_path or project root?
            # data_path is 'data/file_generation/file_generation_mini.json'
            # gt_file is 'data/file_generation/gt/file_generation_gt.xlsx'
            # So it is relative to project root.
            
            gt_file = os.path.abspath(gt_file_rel)
            
            sheet_name = item.get('sheet_name')
            cell_range = item.get('range')
            
            print(f"Evaluating item {i+1}/{len(data)}: {sheet_name}!{cell_range} against {gt_file_rel}")
            
            if not os.path.exists(gt_file):
                print(f"  Warning: GT file not found: {gt_file}")
                res = {'score': 0, 'reason': "GT file not found"}
                results.append({**item, **res})
                continue
                
            try:
                wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
                
                success, msg = cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range)
                
                score = 1 if success else 0
                res = {'score': score, 'reason': msg}
                print(f"  Score: {score}, Reason: {msg}")
                
                wb_gt.close()
                
            except Exception as e:
                print(f"  Error evaluating item: {e}")
                res = {'score': 0, 'reason': str(e)}
            
            results.append({**item, **res})
            total_score += res['score']
            
        wb_proc.close()
        
        print(f"\nFile Generation Evaluation Complete. Total Score: {total_score}/{len(data)}")
        return results
